from openpyxl import Workbook, load_workbook  # 只支援2010以上
from openpyxl.utils import  get_column_letter
#讀取excel

# wb = load_workbook('python_try.xlsx')
# ws = wb['test2']
# wb.create_sheet('nw_test_1')
# wb.save('python_try.xlsx')
# print(wb.sheetnames)

#創建excel

# wb = Workbook()
# ws = wb.active
# ws.title = 'qq'
#
# ws.append([123, 456, 789, 0])
# ws.append([123, 456, 789, 0])
# ws.append([123, 456, 789, 0])
# wb.save("nw_excel.xlsx")

#讀取excel範圍資料
wb = load_workbook("nw_excel.xlsx")
ws = wb.active
for row in range(1, 4):
    for col in range(1, 4):
        char = get_column_letter(col)
        ws[char + str(row)].value = char + str(row)

wb.save("nw_excel.xlsx")